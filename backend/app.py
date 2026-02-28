"""SİTEY-VM Demo - FastAPI Ana Uygulama"""
import os
import sys
import re
import uuid
import io
import json
import time
import html
import hashlib
import threading
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from typing import Optional, List
from collections import defaultdict

from fastapi import FastAPI, Depends, HTTPException, status, Query, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.security import OAuth2PasswordBearer
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func
from pydantic import BaseModel, field_validator

from database import get_db, init_db
from models import User, Vulnerability
from security import (
    create_access_token,
    verify_token,
    verify_password,
    get_password_hash,
)


_login_attempts: dict = defaultdict(list)
MAX_LOGIN_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300
LOGIN_LOCKOUT_SECONDS = 900


def _check_rate_limit(client_ip: str) -> bool:
    """True döndürürse istek engellenmeli."""
    now = time.time()
    attempts = _login_attempts[client_ip]
    _login_attempts[client_ip] = [t for t in attempts if now - t < LOGIN_LOCKOUT_SECONDS]
    attempts = _login_attempts[client_ip]
    recent = [t for t in attempts if now - t < LOGIN_WINDOW_SECONDS]
    return len(recent) >= MAX_LOGIN_ATTEMPTS


def _record_attempt(client_ip: str):
    _login_attempts[client_ip].append(time.time())


_HTML_TAG_RE = re.compile(r"<[^>]+>", re.DOTALL)


def _sanitize(value: str) -> str:
    """HTML tag ve tehlikeli içeriği temizle."""
    if not value:
        return value
    cleaned = value.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    cleaned = cleaned.replace("\x00", "")
    cleaned = html.escape(cleaned, quote=True)
    return cleaned


def _sanitize_dict(d: dict, fields: list) -> dict:
    """Verilen alanları sanitize et."""
    for f in fields:
        if f in d and isinstance(d[f], str):
            d[f] = _sanitize(d[f])
    return d


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; font-src 'self' data:"
        response.headers["Server"] = "SiteyVM"
        return response


_is_production = os.environ.get("SITEYVM_PRODUCTION", "").lower() in ("1", "true", "yes")

app = FastAPI(
    title="SİTEY-VM Demo",
    version="1.4.0-demo",
    docs_url=None if _is_production else "/docs",
    redoc_url=None if _is_production else "/redoc",
    openapi_url=None if _is_production else "/openapi.json",
)

app.add_middleware(SecurityHeadersMiddleware)

_allowed_origins = os.environ.get("SITEYVM_CORS_ORIGINS", "").split(",") if os.environ.get("SITEYVM_CORS_ORIGINS") else []
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins or ["*"],
    allow_credentials=True if _allowed_origins else False,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")


_token_blacklist: set = set()
_user_token_generation: dict = {}


def _blacklist_token(token: str):
    """Token'ı kara listeye ekle."""
    _token_blacklist.add(token)
    if len(_token_blacklist) > 10000:
        _token_blacklist.clear()


_report_timestamps: dict = {}
REPORT_MAX_PER_MINUTE = 5


def _check_report_rate_limit(username: str):
    """Bir kullanıcının dakikada en fazla 5 rapor üretmesine izin ver."""
    now = time.time()
    timestamps = _report_timestamps.get(username, [])
    timestamps = [t for t in timestamps if now - t < 60]
    _report_timestamps[username] = timestamps
    if len(timestamps) >= REPORT_MAX_PER_MINUTE:
        return True
    timestamps.append(now)
    return False


def _invalidate_user_tokens(username: str):
    """Bir kullanıcının tüm token'larını geçersiz kıl (generation artır)."""
    _user_token_generation[username] = _user_token_generation.get(username, 0) + 1


def _get_user_generation(username: str) -> int:
    return _user_token_generation.get(username, 0)


@app.on_event("startup")
def on_startup():
    init_db()


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    if token in _token_blacklist:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Token geçersiz (oturum kapatılmış)")
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Geçersiz token")
    username = payload.get("sub")
    token_gen = payload.get("gen", 0)
    if token_gen < _get_user_generation(username):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Token geçersiz (şifre değiştirilmiş)")
    user = db.query(User).filter_by(username=username).first()
    if not user:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Kullanıcı bulunamadı")
    if not user.is_active:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Hesap devre dışı")
    return user


class LoginRequest(BaseModel):
    username: str
    password: str

    @field_validator("username", "password")
    @classmethod
    def max_length(cls, v):
        if len(v) > 200:
            raise ValueError("Alan en fazla 200 karakter olabilir")
        return v

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class ManualVulnCreate(BaseModel):
    name: str
    cve: Optional[str] = ""
    risk: str = "Medium"
    description: Optional[str] = ""
    solution: Optional[str] = ""
    target_ip: Optional[str] = ""
    port: Optional[str] = ""
    service: Optional[str] = ""
    cvss_score: Optional[str] = ""

    @field_validator("name")
    @classmethod
    def name_max_length(cls, v):
        if len(v) > 500:
            raise ValueError("Zafiyet adı en fazla 500 karakter olabilir")
        return v

    @field_validator("description", "solution")
    @classmethod
    def text_max_length(cls, v):
        if v and len(v) > 5000:
            raise ValueError("Alan en fazla 5000 karakter olabilir")
        return v

    @field_validator("cve", "target_ip", "port", "service", "cvss_score")
    @classmethod
    def short_field_max_length(cls, v):
        if v and len(v) > 200:
            raise ValueError("Alan en fazla 200 karakter olabilir")
        return v

class ManualVulnUpdate(BaseModel):
    name: Optional[str] = None
    cve: Optional[str] = None
    risk: Optional[str] = None
    description: Optional[str] = None
    solution: Optional[str] = None
    target_ip: Optional[str] = None
    port: Optional[str] = None
    service: Optional[str] = None
    cvss_score: Optional[str] = None

    @field_validator("name")
    @classmethod
    def name_max_length(cls, v):
        if v is not None and len(v) > 500:
            raise ValueError("Zafiyet adı en fazla 500 karakter olabilir")
        return v

    @field_validator("description", "solution")
    @classmethod
    def text_max_length(cls, v):
        if v and len(v) > 5000:
            raise ValueError("Alan en fazla 5000 karakter olabilir")
        return v

    @field_validator("cve", "target_ip", "port", "service", "cvss_score")
    @classmethod
    def short_field_max_length(cls, v):
        if v and len(v) > 200:
            raise ValueError("Alan en fazla 200 karakter olabilir")
        return v

class VulnStatusUpdate(BaseModel):
    status: str

class BulkIdsRequest(BaseModel):
    ids: List[str]

class BulkStatusRequest(BaseModel):
    ids: List[str]
    status: str


@app.get("/api/system/info")
def system_info():
    """Sunucu IP ve versiyon bilgisi. LAN üzerinden erişim URL'sini döndürür."""
    import socket as _socket
    ips = set()
    try:
        s = _socket.socket(_socket.AF_INET, _socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ips.add(s.getsockname()[0])
        s.close()
    except Exception:
        pass
    try:
        hostname = _socket.gethostname()
        for info in _socket.getaddrinfo(hostname, None, _socket.AF_INET):
            ip = info[4][0]
            if not ip.startswith("127."):
                ips.add(ip)
    except Exception:
        pass
    if not ips:
        ips.add("127.0.0.1")
    ip_list = sorted(ips)
    primary = next((ip for ip in ip_list if ip.startswith(("192.168.", "10."))), ip_list[0])
    return {
        "version": "1.4.0-demo",
        "hostname": _socket.gethostname(),
        "primary_ip": primary,
        "all_ips": ip_list,
        "port": 5000,
        "access_url": f"http://{primary}:5000",
    }


@app.post("/api/auth/login")
def login(body: LoginRequest, request: Request, db: Session = Depends(get_db)):
    client_ip = request.client.host if request.client else "unknown"
    if _check_rate_limit(client_ip):
        raise HTTPException(
            status.HTTP_429_TOO_MANY_REQUESTS,
            "Çok fazla başarısız giriş denemesi. 15 dakika sonra tekrar deneyin.",
            headers={"Retry-After": str(LOGIN_LOCKOUT_SECONDS)},
        )

    user = db.query(User).filter_by(username=body.username).first()
    if not user or not verify_password(body.password, user.password_hash):
        _record_attempt(client_ip)
        raise HTTPException(401, "Kullanıcı adı veya şifre hatalı")
    if not user.is_active:
        raise HTTPException(403, "Hesap pasif durumda")
    _login_attempts.pop(client_ip, None)
    user.last_login = datetime.now(timezone.utc)
    db.commit()
    gen = _get_user_generation(user.username)
    token = create_access_token({"sub": user.username, "role": user.role, "gen": gen})
    return {"access_token": token, "token_type": "bearer"}


@app.get("/api/auth/users/me")
def get_me(user: User = Depends(get_current_user)):
    return {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "is_active": user.is_active,
        "first_name": "",
        "last_name": "",
        "title": "Demo Kullanıcı",
    }


@app.get("/api/auth/users/me/language")
def get_language(user: User = Depends(get_current_user)):
    return {"language": "tr"}


@app.patch("/api/auth/users/me/language")
def set_language(user: User = Depends(get_current_user)):
    return {"ok": True}


@app.post("/api/auth/change-password")
def change_password(
    body: PasswordChange,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not verify_password(body.current_password, user.password_hash):
        raise HTTPException(400, "Mevcut şifre hatalı")
    new_pw = body.new_password
    if len(new_pw) < 8:
        raise HTTPException(400, "Şifre en az 8 karakter olmalıdır")
    if not re.search(r"[A-Z]", new_pw):
        raise HTTPException(400, "Şifre en az bir büyük harf içermelidir")
    if not re.search(r"[a-z]", new_pw):
        raise HTTPException(400, "Şifre en az bir küçük harf içermelidir")
    if not re.search(r"[0-9]", new_pw):
        raise HTTPException(400, "Şifre en az bir rakam içermelidir")
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", new_pw):
        raise HTTPException(400, "Şifre en az bir özel karakter içermelidir")
    user.password_hash = get_password_hash(new_pw)
    db.commit()
    _invalidate_user_tokens(user.username)
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        _blacklist_token(auth_header[7:])
    return {"ok": True, "message": "Şifre başarıyla değiştirildi. Lütfen yeniden giriş yapın."}


@app.post("/api/auth/logout")
def logout(request: Request, user: User = Depends(get_current_user)):
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        _blacklist_token(auth_header[7:])
    return {"ok": True, "message": "Oturum kapatıldı"}


@app.get("/api/license/status")
def license_status(user: User = Depends(get_current_user)):
    return {
        "valid": True,
        "plan": "Demo (Ücretsiz)",
        "expire": "Süresiz",
        "ip_limit": 16,
        "used_ip_count": 0,
        "features": ["dashboard", "vulnerability_list", "vulnerability_detail", "reports", "openvas_import", "manual_vuln_add"],
    }


@app.get("/api/vuln/dashboard_stats")
def dashboard_stats(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    vulns = db.query(Vulnerability).all()
    risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
    status_counts = {"open": 0, "in_progress": 0, "resolved": 0, "false_positive": 0}
    scanners = {}
    for v in vulns:
        risk_counts[v.risk] = risk_counts.get(v.risk, 0) + 1
        status_counts[v.status] = status_counts.get(v.status, 0) + 1
        scanners[v.scanner or "Manual"] = scanners.get(v.scanner or "Manual", 0) + 1
    return {
        "total": len(vulns),
        "by_risk": risk_counts,
        "by_status": status_counts,
        "by_scanner": scanners,
        "recent_vulns": [_vuln_dict(v) for v in vulns[:5]],
    }


@app.get("/api/vuln/dashboard_advanced_stats")
def dashboard_advanced_stats(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    vulns = db.query(Vulnerability).all()
    risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
    for v in vulns:
        risk_counts[v.risk] = risk_counts.get(v.risk, 0) + 1
    return {
        "total_vulns": len(vulns),
        "critical": risk_counts["Critical"],
        "high": risk_counts["High"],
        "medium": risk_counts["Medium"],
        "low": risk_counts["Low"],
        "info": risk_counts["Info"],
        "open_vulns": sum(1 for v in vulns if v.status == "open"),
        "resolved_vulns": sum(1 for v in vulns if v.status == "resolved"),
        "in_progress_vulns": sum(1 for v in vulns if v.status == "in_progress"),
        "resolution_rate": 0,
        "avg_resolution_days": 0,
        "trend_data": [],
        "top_ips": _top_ips(vulns),
        "assets_count": len(set(v.target_ip for v in vulns if v.target_ip)),
    }


@app.get("/api/scan/dashboard_stats")
def scan_dashboard_stats(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return dashboard_stats(user=user, db=db)


@app.get("/api/scan/dashboard_advanced_stats")
def scan_dashboard_advanced_stats(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return dashboard_advanced_stats(user=user, db=db)


import os as _os

def _find_logo():
    """LOGO.png'yi hem dev hem EXE modunda bul."""
    candidates = [
        _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "LOGO.png"),
        _os.path.join(getattr(sys, '_MEIPASS', ''), "backend", "LOGO.png"),
        _os.path.join(_os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else '', "backend", "LOGO.png"),
    ]
    for p in candidates:
        if _os.path.isfile(p):
            return p
    return candidates[0]

_LOGO_PATH = _find_logo()

_FONT_REGISTERED = False
def _register_turkish_fonts():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    import sys

    # --- reportlab paketinin kendi içindeki DejaVu fontlarını da ara ---
    _rl_font_dir = None
    try:
        import reportlab
        _rl_font_dir = _os.path.join(_os.path.dirname(reportlab.__file__), "fonts")
    except Exception:
        pass

    _font_dirs = [
        _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "fonts"),
        "/usr/share/fonts/truetype/dejavu",
    ]
    # reportlab'ın kendi font dizinini ekle
    if _rl_font_dir:
        _font_dirs.append(_rl_font_dir)
    # Windows Fonts dizini
    _font_dirs.append(
        _os.path.join(_os.environ.get("WINDIR", "C:\\Windows"), "Fonts")
    )

    # ---- 1) Önce DejaVu fontlarını ara ----
    _dejavu = _dejavu_b = _dejavu_o = _dejavu_bo = None
    for _fd in _font_dirs:
        _candidate = _os.path.join(_fd, "DejaVuSans.ttf")
        if _os.path.exists(_candidate):
            _dejavu = _candidate
            _dejavu_b = _os.path.join(_fd, "DejaVuSans-Bold.ttf")
            _dejavu_o = _os.path.join(_fd, "DejaVuSans-Oblique.ttf")
            _dejavu_bo = _os.path.join(_fd, "DejaVuSans-BoldOblique.ttf")
            break

    if _dejavu and _os.path.exists(_dejavu):
        pdfmetrics.registerFont(TTFont("DejaVu", _dejavu))
        if _os.path.exists(_dejavu_b):
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", _dejavu_b))
        else:
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", _dejavu))
        if _os.path.exists(_dejavu_o):
            pdfmetrics.registerFont(TTFont("DejaVu-Italic", _dejavu_o))
        else:
            pdfmetrics.registerFont(TTFont("DejaVu-Italic", _dejavu))
        if _os.path.exists(_dejavu_bo):
            pdfmetrics.registerFont(TTFont("DejaVu-BoldItalic", _dejavu_bo))
        else:
            pdfmetrics.registerFont(TTFont("DejaVu-BoldItalic", _dejavu))
        registerFontFamily("DejaVu", normal="DejaVu", bold="DejaVu-Bold",
                           italic="DejaVu-Italic", boldItalic="DejaVu-BoldItalic")
        _FONT_REGISTERED = True
        return

    # ---- 2) DejaVu bulunamadı — Windows yerleşik fontlarını dene ----
    if sys.platform == "win32":
        _win_fonts_dir = _os.path.join(
            _os.environ.get("WINDIR", "C:\\Windows"), "Fonts"
        )
        # Tercih sırası: Segoe UI > Arial > Tahoma > Verdana > Calibri
        _win_font_sets = [
            {
                "name": "SegoeUI",
                "normal": "segoeui.ttf",
                "bold": "segoeuib.ttf",
                "italic": "segoeuii.ttf",
                "bolditalic": "segoeuiz.ttf",
            },
            {
                "name": "Arial",
                "normal": "arial.ttf",
                "bold": "arialbd.ttf",
                "italic": "ariali.ttf",
                "bolditalic": "arialbi.ttf",
            },
            {
                "name": "Tahoma",
                "normal": "tahoma.ttf",
                "bold": "tahomabd.ttf",
                "italic": "tahoma.ttf",
                "bolditalic": "tahomabd.ttf",
            },
            {
                "name": "Verdana",
                "normal": "verdana.ttf",
                "bold": "verdanab.ttf",
                "italic": "verdanai.ttf",
                "bolditalic": "verdanaz.ttf",
            },
            {
                "name": "Calibri",
                "normal": "calibri.ttf",
                "bold": "calibrib.ttf",
                "italic": "calibrii.ttf",
                "bolditalic": "calibriz.ttf",
            },
        ]
        for fset in _win_font_sets:
            _normal = _os.path.join(_win_fonts_dir, fset["normal"])
            if not _os.path.exists(_normal):
                continue
            _bold = _os.path.join(_win_fonts_dir, fset["bold"])
            _italic = _os.path.join(_win_fonts_dir, fset["italic"])
            _bolditalic = _os.path.join(_win_fonts_dir, fset["bolditalic"])
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", _normal))
                pdfmetrics.registerFont(TTFont("DejaVu-Bold",
                    _bold if _os.path.exists(_bold) else _normal))
                pdfmetrics.registerFont(TTFont("DejaVu-Italic",
                    _italic if _os.path.exists(_italic) else _normal))
                pdfmetrics.registerFont(TTFont("DejaVu-BoldItalic",
                    _bolditalic if _os.path.exists(_bolditalic) else _normal))
                registerFontFamily("DejaVu", normal="DejaVu", bold="DejaVu-Bold",
                                   italic="DejaVu-Italic", boldItalic="DejaVu-BoldItalic")
                _FONT_REGISTERED = True
                return
            except Exception:
                continue

@app.post("/api/vuln/report/pdf")
def generate_pdf_report(
    body: dict = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if _check_report_rate_limit(user.username):
        raise HTTPException(429, "Çok fazla rapor isteği. Lütfen 1 dakika bekleyin.")
    _register_turkish_fonts()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
        HRFlowable, PageBreak, KeepTogether
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    FONT = "DejaVu" if _FONT_REGISTERED else "Helvetica"
    FONT_B = "DejaVu-Bold" if _FONT_REGISTERED else "Helvetica-Bold"
    FONT_I = "DejaVu-Italic" if _FONT_REGISTERED else "Helvetica-Oblique"

    selected_ids = (body or {}).get("ids", [])
    if selected_ids:
        vulns = db.query(Vulnerability).filter(Vulnerability.id.in_(selected_ids)).order_by(Vulnerability.risk).all()
    else:
        vulns = db.query(Vulnerability).order_by(Vulnerability.risk).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20*mm, bottomMargin=25*mm,
        leftMargin=18*mm, rightMargin=18*mm,
    )

    brand_blue = rl_colors.HexColor("#2563eb")
    brand_dark = rl_colors.HexColor("#0f172a")
    brand_gray = rl_colors.HexColor("#64748b")
    brand_light = rl_colors.HexColor("#f8fafc")
    brand_orange = rl_colors.HexColor("#f97316")
    white = rl_colors.white

    s_cover_title = ParagraphStyle("CoverTitle", fontName=FONT_B, fontSize=28, textColor=brand_dark, alignment=TA_CENTER, spaceAfter=6, leading=34)
    s_cover_sub = ParagraphStyle("CoverSub", fontName=FONT, fontSize=14, textColor=brand_blue, alignment=TA_CENTER, spaceAfter=4)
    s_cover_info = ParagraphStyle("CoverInfo", fontName=FONT, fontSize=11, textColor=brand_gray, alignment=TA_CENTER, spaceAfter=3)
    s_cover_demo = ParagraphStyle("CoverDemo", fontName=FONT_I, fontSize=10, textColor=brand_orange, alignment=TA_CENTER, spaceBefore=12, spaceAfter=2)

    s_title = ParagraphStyle("RTitle", fontName=FONT_B, fontSize=18, textColor=brand_dark, spaceAfter=4, alignment=TA_LEFT)
    s_sub = ParagraphStyle("RSub", fontName=FONT, fontSize=10, textColor=brand_gray, spaceAfter=2)
    s_heading = ParagraphStyle("RHeading", fontName=FONT_B, fontSize=13, textColor=brand_blue, spaceBefore=18, spaceAfter=8)
    s_body = ParagraphStyle("RBody", fontName=FONT, fontSize=9, leading=13, textColor=brand_dark)
    s_footer = ParagraphStyle("RFooter", fontName=FONT, fontSize=8, textColor=brand_gray, alignment=TA_CENTER)
    s_demo = ParagraphStyle("RDemo", fontName=FONT_I, fontSize=9, textColor=brand_orange, alignment=TA_CENTER, spaceBefore=6)

    s_rec_title = ParagraphStyle("RecTitle", fontName=FONT_B, fontSize=16, textColor=brand_dark, alignment=TA_CENTER, spaceAfter=8, spaceBefore=4)
    s_rec_heading = ParagraphStyle("RecHeading", fontName=FONT_B, fontSize=11, textColor=brand_blue, spaceBefore=12, spaceAfter=4)
    s_rec_body = ParagraphStyle("RecBody", fontName=FONT, fontSize=9, leading=14, textColor=rl_colors.HexColor("#334155"), spaceAfter=2)
    s_rec_bullet = ParagraphStyle("RecBullet", fontName=FONT, fontSize=9, leading=14, textColor=rl_colors.HexColor("#334155"), leftIndent=16, spaceAfter=2)
    s_promo_title = ParagraphStyle("PromoTitle", fontName=FONT_B, fontSize=14, textColor=brand_blue, alignment=TA_CENTER, spaceBefore=20, spaceAfter=6)
    s_promo_body = ParagraphStyle("PromoBody", fontName=FONT, fontSize=9, leading=14, textColor=rl_colors.HexColor("#334155"), alignment=TA_CENTER, spaceAfter=2)
    s_promo_url = ParagraphStyle("PromoURL", fontName=FONT_B, fontSize=12, textColor=brand_blue, alignment=TA_CENTER, spaceBefore=8, spaceAfter=2)

    elements = []

    risk_labels = {"Critical": "Kritik", "High": "Yüksek", "Medium": "Orta", "Low": "Düşük", "Info": "Bilgi"}
    risk_colors_map = {
        "Critical": rl_colors.HexColor("#dc2626"),
        "High": rl_colors.HexColor("#ea580c"),
        "Medium": rl_colors.HexColor("#d97706"),
        "Low": rl_colors.HexColor("#2563eb"),
        "Info": rl_colors.HexColor("#64748b"),
    }
    risk_bg_map = {
        "Critical": rl_colors.HexColor("#fef2f2"),
        "High": rl_colors.HexColor("#fff7ed"),
        "Medium": rl_colors.HexColor("#fefce8"),
        "Low": rl_colors.HexColor("#eff6ff"),
        "Info": rl_colors.HexColor("#f9fafb"),
    }
    risk_count = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
    for v in vulns:
        risk_count[v.risk] = risk_count.get(v.risk, 0) + 1
    status_tr = {"open": "Açık", "in_progress": "Devam", "resolved": "Çözüldü", "false_positive": "YP"}

    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    elements.append(Spacer(1, 60))

    if _os.path.exists(_LOGO_PATH):
        logo = Image(_LOGO_PATH, width=180, height=102)
        logo.hAlign = "CENTER"
        elements.append(logo)
    elements.append(Spacer(1, 24))

    elements.append(Paragraph("SİTEY-VM", s_cover_title))
    elements.append(Paragraph("Kurumsal Zafiyet Yönetim Platformu", s_cover_sub))
    elements.append(Spacer(1, 20))

    elements.append(HRFlowable(width="60%", thickness=2, color=brand_blue, spaceBefore=8, spaceAfter=16, hAlign="CENTER"))

    elements.append(Paragraph("Zafiyet Analiz Raporu", ParagraphStyle("CT2", fontName=FONT_B, fontSize=18, textColor=brand_dark, alignment=TA_CENTER, spaceAfter=16)))
    elements.append(Spacer(1, 12))

    cover_data = [
        [Paragraph("<b>Rapor Tarihi</b>", s_body), Paragraph(now_str, s_body)],
        [Paragraph("<b>Hazırlayan</b>", s_body), Paragraph(user.username, s_body)],
        [Paragraph("<b>Toplam Zafiyet</b>", s_body), Paragraph(str(len(vulns)), s_body)],
        [Paragraph("<b>Kritik + Yüksek</b>", s_body), Paragraph(str(risk_count["Critical"] + risk_count["High"]), s_body)],
        [Paragraph("<b>Rapor Türü</b>", s_body), Paragraph("Seçili Zafiyetler" if selected_ids else "Tüm Zafiyetler", s_body)],
    ]
    cover_tbl = Table(cover_data, colWidths=[120, 200], hAlign="CENTER")
    cover_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#e2e8f0")),
        ("BACKGROUND", (0, 0), (0, -1), rl_colors.HexColor("#f1f5f9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(cover_tbl)
    elements.append(Spacer(1, 30))

    elements.append(Paragraph("Bu rapor SİTEY-VM Demo Sürümü ile oluşturulmuştur.", s_cover_demo))
    elements.append(Paragraph("Kurumsal sürüm için: siteyvm.com", s_cover_info))
    elements.append(Spacer(1, 20))

    cover_risk = [[Paragraph(f"<b>{risk_labels[k]}</b>", ParagraphStyle("CRK", fontName=FONT_B, fontSize=10, textColor=risk_colors_map[k], alignment=TA_CENTER)),
                    Paragraph(str(risk_count[k]), ParagraphStyle("CRV", fontName=FONT_B, fontSize=14, textColor=risk_colors_map[k], alignment=TA_CENTER))]
                   for k in ["Critical", "High", "Medium", "Low", "Info"]]
    risk_labels_row = [r[0] for r in cover_risk]
    risk_counts_row = [r[1] for r in cover_risk]
    risk_tbl = Table([risk_counts_row, risk_labels_row], colWidths=[65]*5, hAlign="CENTER")
    risk_tbl.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (0, -1), risk_bg_map["Critical"]),
        ("BACKGROUND", (1, 0), (1, -1), risk_bg_map["High"]),
        ("BACKGROUND", (2, 0), (2, -1), risk_bg_map["Medium"]),
        ("BACKGROUND", (3, 0), (3, -1), risk_bg_map["Low"]),
        ("BACKGROUND", (4, 0), (4, -1), risk_bg_map["Info"]),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(risk_tbl)

    elements.append(PageBreak())


    header_data = []
    if _os.path.exists(_LOGO_PATH):
        small_logo = Image(_LOGO_PATH, width=32, height=18)
        header_data = [[small_logo, Paragraph("SİTEY-VM — Zafiyet Analiz Raporu", s_title)]]
    else:
        header_data = [["", Paragraph("SİTEY-VM — Zafiyet Analiz Raporu", s_title)]]
    header_table = Table(header_data, colWidths=[40, 450])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Paragraph(f"Oluşturulma: {now_str}  |  Zafiyet: {len(vulns)}  |  Hazırlayan: {user.username}", s_sub))
    elements.append(HRFlowable(width="100%", thickness=1, color=brand_blue, spaceBefore=6, spaceAfter=12))

    elements.append(Paragraph("Risk Dağılımı", s_heading))
    summary_data = [[
        Paragraph("<b>Seviye</b>", ParagraphStyle("sh", fontName=FONT_B, fontSize=9, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Türkçe</b>", ParagraphStyle("sh2", fontName=FONT_B, fontSize=9, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Adet</b>", ParagraphStyle("sh3", fontName=FONT_B, fontSize=9, textColor=white, alignment=TA_CENTER)),
        Paragraph("<b>Oran</b>", ParagraphStyle("sh4", fontName=FONT_B, fontSize=9, textColor=white, alignment=TA_CENTER)),
    ]]
    for k in ["Critical", "High", "Medium", "Low", "Info"]:
        c = risk_count[k]
        pct = f"%{round(c/len(vulns)*100)}" if vulns else "%0"
        summary_data.append([
            Paragraph(f"<b>{k}</b>", ParagraphStyle(f"sr_{k}", fontName=FONT_B, fontSize=9, textColor=risk_colors_map[k])),
            Paragraph(risk_labels[k], ParagraphStyle(f"srt_{k}", fontName=FONT, fontSize=9, textColor=brand_dark)),
            Paragraph(f"<b>{c}</b>", ParagraphStyle(f"src_{k}", fontName=FONT_B, fontSize=9, textColor=brand_dark, alignment=TA_CENTER)),
            Paragraph(pct, ParagraphStyle(f"srp_{k}", fontName=FONT, fontSize=9, textColor=brand_dark, alignment=TA_CENTER)),
        ])
    summary_data.append([
        Paragraph("", s_body),
        Paragraph("<b>TOPLAM</b>", ParagraphStyle("st", fontName=FONT_B, fontSize=9, textColor=brand_dark)),
        Paragraph(f"<b>{len(vulns)}</b>", ParagraphStyle("stc", fontName=FONT_B, fontSize=9, textColor=brand_dark, alignment=TA_CENTER)),
        Paragraph("<b>%100</b>", ParagraphStyle("stp", fontName=FONT_B, fontSize=9, textColor=brand_dark, alignment=TA_CENTER)),
    ])

    st = Table(summary_data, colWidths=[80, 80, 60, 60])
    st_style = [
        ("BACKGROUND", (0, 0), (-1, 0), brand_dark),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#e2e8f0")),
        ("BACKGROUND", (0, -1), (-1, -1), rl_colors.HexColor("#f1f5f9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for idx, k in enumerate(["Critical", "High", "Medium", "Low", "Info"], 1):
        st_style.append(("BACKGROUND", (0, idx), (0, idx), risk_bg_map[k]))
    st.setStyle(TableStyle(st_style))
    elements.append(st)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Zafiyet Detayları", s_heading))

    hdr_style = ParagraphStyle("dh", fontName=FONT_B, fontSize=7, textColor=white, alignment=TA_CENTER)
    detail_header = [
        Paragraph("#", hdr_style),
        Paragraph("Zafiyet Adı", hdr_style),
        Paragraph("CVE", hdr_style),
        Paragraph("Seviye", hdr_style),
        Paragraph("CVSS", hdr_style),
        Paragraph("IP", hdr_style),
        Paragraph("Port", hdr_style),
        Paragraph("Durum", hdr_style),
    ]
    detail_data = [detail_header]

    cell_style = ParagraphStyle("dc", fontName=FONT, fontSize=7, leading=9, textColor=brand_dark)
    cell_center = ParagraphStyle("dcc", fontName=FONT, fontSize=7, leading=9, textColor=brand_dark, alignment=TA_CENTER)

    for i, v in enumerate(vulns, 1):
        name = (v.name or "İsimsiz")[:50]
        risk_cell_style = ParagraphStyle(f"dr_{i}", fontName=FONT_B, fontSize=7, textColor=risk_colors_map.get(v.risk, brand_dark), alignment=TA_CENTER)
        detail_data.append([
            Paragraph(str(i), cell_center),
            Paragraph(name, cell_style),
            Paragraph(v.cve or "—", cell_style),
            Paragraph(risk_labels.get(v.risk, v.risk or "—"), risk_cell_style),
            Paragraph(str(v.cvss_score or "—"), cell_center),
            Paragraph(v.target_ip or "—", cell_style),
            Paragraph(v.port or "—", cell_center),
            Paragraph(status_tr.get(v.status, v.status or "—"), cell_center),
        ])

    dt = Table(detail_data, colWidths=[22, 155, 75, 48, 32, 72, 32, 50], repeatRows=1)
    dt_style = [
        ("BACKGROUND", (0, 0), (-1, 0), brand_blue),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, rl_colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, v in enumerate(vulns, 1):
        bg = risk_bg_map.get(v.risk)
        if bg and v.risk in ("Critical", "High"):
            dt_style.append(("BACKGROUND", (3, i), (3, i), bg))
    dt.setStyle(TableStyle(dt_style))
    elements.append(dt)

    elements.append(PageBreak())

    elements.append(Paragraph("Zafiyet Yönetimi Güçlendirme Önerileri", s_rec_title))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=brand_blue, spaceBefore=4, spaceAfter=14, hAlign="CENTER"))

    if risk_count["Critical"] > 0:
        elements.append(Paragraph("Kritik Seviye Zafiyetler", s_rec_heading))
        elements.append(Paragraph(f"Raporunuzda <b>{risk_count['Critical']}</b> adet kritik seviye zafiyet tespit edilmiştir. Bu zafiyetler, saldırganların sisteme tam yetkili erişim sağlamasına olanak tanıyabilir.", s_rec_body))
        elements.append(Paragraph("• Kritik zafiyetler için derhal yama planı oluşturun ve 24-48 saat içinde aksiyon alın.", s_rec_bullet))
        elements.append(Paragraph("• Etkilenen sistemleri ağ segmentasyonu ile izole edin.", s_rec_bullet))
        elements.append(Paragraph("• Olay müdahale ekibinizi bilgilendirin ve IOC kontrolü yapın.", s_rec_bullet))

    if risk_count["High"] > 0:
        elements.append(Paragraph("Yüksek Seviye Zafiyetler", s_rec_heading))
        elements.append(Paragraph(f"Raporunuzda <b>{risk_count['High']}</b> adet yüksek seviye zafiyet bulunmaktadır. Bu zafiyetler yetki yükseltme, veri sızıntısı veya servis kesintisine yol açabilir.", s_rec_body))
        elements.append(Paragraph("• 1 hafta içinde yama veya geçici çözüm (workaround) uygulayın.", s_rec_bullet))
        elements.append(Paragraph("• Güvenlik duvarı kurallarını gözden geçirip sıkılaştırın.", s_rec_bullet))
        elements.append(Paragraph("• Etkilenen servislerde erişim loglarını analiz edin.", s_rec_bullet))

    if risk_count["Medium"] > 0:
        elements.append(Paragraph("Orta Seviye Zafiyetler", s_rec_heading))
        elements.append(Paragraph(f"<b>{risk_count['Medium']}</b> adet orta seviye zafiyet tespit edilmiştir.", s_rec_body))
        elements.append(Paragraph("• Aylık bakım penceresinde bu zafiyetleri giderin.", s_rec_bullet))
        elements.append(Paragraph("• Varsayılan yapılandırmaları kontrol edip sertleştirin.", s_rec_bullet))

    elements.append(Paragraph("Genel Güvenlik Önerileri", s_rec_heading))
    general_recs = [
        "• Düzenli zafiyet taraması yapın — en az ayda bir kez tüm varlıkları tarayın.",
        "• Yama yönetimi sürecinizi otomatikleştirin ve SLA'lar tanımlayın.",
        "• Ağ segmentasyonu uygulayarak kritik sistemleri izole edin.",
        "• Çok faktörlü kimlik doğrulama (MFA) tüm yönetim arayüzlerinde aktif olsun.",
        "• Güvenlik bilinci eğitimleri ile personel farkındalığını artırın.",
        "• Olay müdahale planınızı güncel tutun ve yılda en az iki kez tatbikat yapın.",
        "• Sızma testi (pentest) ile zafiyet taramalarını doğrulayın.",
        "• Log yönetimi ile merkezi izleme yapın.",
    ]
    for rec in general_recs:
        elements.append(Paragraph(rec, s_rec_bullet))

    elements.append(Spacer(1, 16))
    elements.append(HRFlowable(width="80%", thickness=1, color=rl_colors.HexColor("#e2e8f0"), spaceBefore=8, spaceAfter=12, hAlign="CENTER"))

    elements.append(Paragraph("SİTEY-VM ile Zafiyet Yönetiminizi Güçlendirin", s_promo_title))

    promo_features = [
        "Çoklu tarayıcı entegrasyonu (OpenVAS, Nessus, Qualys, Acunetix, Nuclei ve dahası)",
        "Yapay zeka destekli zafiyet analizi ve önceliklendirme",
        "Otomatik yama takibi ve SLA yönetimi",
        "Rol tabanlı erişim kontrolü (RBAC) ile ekip yönetimi",
        "İş akışı motoru — zafiyet keşiften çözüme kadar tam izlenebilirlik",
        "Saldırı yüzeyi yönetimi (ASM) ve sürekli varlık keşfi",
        "Kapsamlı raporlama: PDF, Excel, yönetici özeti, teknik detay",
    ]
    for feat in promo_features:
        elements.append(Paragraph(f"✓  {feat}", s_rec_bullet))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("siteyvm.com", s_promo_url))
    elements.append(Paragraph("satis@siteyvm.com  |  Kurumsal lisans ve demo talebi için bize ulaşın.", s_promo_body))

    elements.append(Spacer(1, 24))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.HexColor("#e2e8f0"), spaceBefore=10, spaceAfter=8))
    elements.append(Paragraph("Bu rapor SİTEY-VM Demo Sürümü v1.4.0 ile oluşturulmuştur.", s_footer))
    elements.append(Paragraph("Kurumsal lisans ile tam özellikli raporlama, AI analiz ve daha fazlasına erişin.", s_footer))
    elements.append(Paragraph(f"© 2025-2026 SİTEY Siber Güvenlik  |  siteyvm.com  |  Rapor Tarihi: {now_str}", s_footer))

    doc.build(elements)
    buf.seek(0)
    filename = f"siteyvm_demo_rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/api/vuln/report/excel")
def generate_excel_report(
    body: dict = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if _check_report_rate_limit(user.username):
        raise HTTPException(429, "Çok fazla rapor isteği. Lütfen 1 dakika bekleyin.")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    selected_ids = (body or {}).get("ids", [])
    if selected_ids:
        vulns = db.query(Vulnerability).filter(Vulnerability.id.in_(selected_ids)).order_by(Vulnerability.risk).all()
    else:
        vulns = db.query(Vulnerability).order_by(Vulnerability.risk).all()

    wb = Workbook()

    ws_cover = wb.active
    ws_cover.title = "Rapor Bilgileri"
    ws_cover.sheet_properties.tabColor = "2563eb"

    brand_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
    brand_font = Font(color="FFFFFF", bold=True, size=14)
    info_font = Font(color="334155", size=11)
    demo_font = Font(color="f97316", size=10, italic=True)
    link_font = Font(color="2563eb", size=11, underline="single")
    thin_border = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )

    ws_cover.merge_cells("A1:D1")
    c = ws_cover["A1"]
    c.value = "SİTEY-VM — Zafiyet Raporu"
    c.fill = brand_fill
    c.font = brand_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 40

    info_rows = [
        ("Ürün", "SİTEY-VM Kurumsal Zafiyet Yönetim Platformu"),
        ("Sürüm", "Demo Sürümü v1.4.0"),
        ("Oluşturan", user.username),
        ("Tarih", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Toplam Zafiyet", str(len(vulns))),
        ("Web Sitesi", "https://siteyvm.com"),
        ("İletişim", "satis@siteyvm.com"),
        ("", ""),
        ("⚠ UYARI", "Bu rapor SİTEY-VM Demo sürümü ile oluşturulmuştur."),
        ("", "Tam kapsamlı raporlama için kurumsal lisans gereklidir."),
        ("", "Detaylı bilgi: siteyvm.com"),
    ]
    for i, (label, value) in enumerate(info_rows, 3):
        ws_cover.cell(row=i, column=1, value=label).font = Font(bold=True, color="334155", size=11)
        cell = ws_cover.cell(row=i, column=2, value=value)
        if "siteyvm.com" in value:
            cell.font = link_font
        elif "UYARI" in label or "Demo" in value or "lisans" in value:
            cell.font = demo_font
        else:
            cell.font = info_font

    ws_cover.column_dimensions["A"].width = 20
    ws_cover.column_dimensions["B"].width = 55

    ws_sum = wb.create_sheet("Risk Özeti")
    ws_sum.sheet_properties.tabColor = "10b981"
    sum_headers = ["Seviye", "Türkçe", "Adet", "Oran"]
    sum_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    sum_font = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=col, value=h)
        c.fill = sum_fill
        c.font = sum_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    risk_labels_map = {"Critical": "Kritik", "High": "Yüksek", "Medium": "Orta", "Low": "Düşük", "Info": "Bilgi"}
    risk_fill_map = {
        "Critical": PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid"),
        "High": PatternFill(start_color="fff7ed", end_color="fff7ed", fill_type="solid"),
        "Medium": PatternFill(start_color="fefce8", end_color="fefce8", fill_type="solid"),
        "Low": PatternFill(start_color="eff6ff", end_color="eff6ff", fill_type="solid"),
        "Info": PatternFill(start_color="f9fafb", end_color="f9fafb", fill_type="solid"),
    }
    risk_font_map = {
        "Critical": Font(color="dc2626", bold=True, size=11),
        "High": Font(color="ea580c", bold=True, size=11),
        "Medium": Font(color="d97706", bold=True, size=11),
        "Low": Font(color="2563eb", bold=True, size=11),
        "Info": Font(color="64748b", bold=True, size=11),
    }
    risk_count = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
    for v in vulns:
        risk_count[v.risk] = risk_count.get(v.risk, 0) + 1

    for idx, k in enumerate(["Critical", "High", "Medium", "Low", "Info"], 2):
        cnt = risk_count[k]
        pct = f"%{round(cnt/len(vulns)*100)}" if vulns else "0%"
        c1 = ws_sum.cell(row=idx, column=1, value=k)
        c1.fill = risk_fill_map[k]
        c1.font = risk_font_map[k]
        c1.border = thin_border
        ws_sum.cell(row=idx, column=2, value=risk_labels_map[k]).border = thin_border
        c3 = ws_sum.cell(row=idx, column=3, value=cnt)
        c3.alignment = Alignment(horizontal="center")
        c3.border = thin_border
        c4 = ws_sum.cell(row=idx, column=4, value=pct)
        c4.alignment = Alignment(horizontal="center")
        c4.border = thin_border

    total_row = 7
    ws_sum.cell(row=total_row, column=1).border = thin_border
    c_t = ws_sum.cell(row=total_row, column=2, value="TOPLAM")
    c_t.font = Font(bold=True, size=11)
    c_t.border = thin_border
    c_n = ws_sum.cell(row=total_row, column=3, value=len(vulns))
    c_n.font = Font(bold=True, size=11)
    c_n.alignment = Alignment(horizontal="center")
    c_n.border = thin_border
    c_p = ws_sum.cell(row=total_row, column=4, value="%100")
    c_p.font = Font(bold=True, size=11)
    c_p.alignment = Alignment(horizontal="center")
    c_p.border = thin_border

    for c in ["A", "B", "C", "D"]:
        ws_sum.column_dimensions[c].width = 16

    ws = wb.create_sheet("Zafiyet Detayları")
    ws.sheet_properties.tabColor = "ef4444"

    headers = ["#", "Zafiyet Adı", "CVE", "Seviye", "CVSS", "Hedef IP", "Port", "Servis", "Durum", "Tarayıcı", "Açıklama", "Çözüm"]
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 24

    status_tr = {"open": "Açık", "in_progress": "Devam Ediyor", "resolved": "Çözüldü", "false_positive": "Yanlış Pozitif"}
    for row_idx, v in enumerate(vulns, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=v.name or "")
        ws.cell(row=row_idx, column=3, value=v.cve or "")
        risk_cell = ws.cell(row=row_idx, column=4, value=v.risk or "")
        risk_cell.fill = risk_fill_map.get(v.risk, PatternFill())
        risk_cell.font = risk_font_map.get(v.risk, Font())
        risk_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=5, value=v.cvss_score or "").alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=v.target_ip or "")
        ws.cell(row=row_idx, column=7, value=v.port or "").alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=8, value=v.service or "")
        ws.cell(row=row_idx, column=9, value=status_tr.get(v.status, v.status or ""))
        ws.cell(row=row_idx, column=10, value=v.scanner or "")
        ws.cell(row=row_idx, column=11, value=(v.description or "")[:300])
        ws.cell(row=row_idx, column=12, value=(v.solution or "")[:300])
        for c in range(1, 13):
            ws.cell(row=row_idx, column=c).border = thin_border

    widths = [5, 35, 18, 10, 8, 15, 8, 12, 14, 12, 45, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws_note = wb.create_sheet("Öneriler ve Hakkında")
    ws_note.sheet_properties.tabColor = "f97316"

    rec_header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    rec_header_font = Font(color="FFFFFF", bold=True, size=12)
    rec_section_font = Font(color="2563eb", bold=True, size=11)
    rec_body_font = Font(color="334155", size=10)
    rec_bold_font = Font(color="334155", bold=True, size=10)
    promo_fill = PatternFill(start_color="eff6ff", end_color="eff6ff", fill_type="solid")
    promo_font = Font(color="2563eb", bold=True, size=11)
    promo_body_font = Font(color="334155", size=10)
    demo_warn_fill = PatternFill(start_color="fff7ed", end_color="fff7ed", fill_type="solid")

    ws_note.merge_cells("A1:B1")
    c = ws_note["A1"]
    c.value = "Zafiyet Yönetimi Güçlendirme Önerileri"
    c.fill = rec_header_fill
    c.font = rec_header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_note.row_dimensions[1].height = 32

    risk_count_xl = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
    for v in vulns:
        risk_count_xl[v.risk] = risk_count_xl.get(v.risk, 0) + 1

    row = 3
    if risk_count_xl["Critical"] > 0:
        ws_note.cell(row=row, column=1, value="Kritik Seviye Zafiyetler").font = rec_section_font
        row += 1
        ws_note.cell(row=row, column=1, value=f"{risk_count_xl['Critical']} adet kritik zafiyet tespit edilmiştir.").font = rec_bold_font
        row += 1
        for rec in [
            "Derhal yama planı oluşturun, 24-48 saat içinde aksiyon alın.",
            "Etkilenen sistemleri ağ segmentasyonu ile izole edin.",
            "Olay müdahale ekibini bilgilendirin ve IOC kontrolü yapın.",
        ]:
            ws_note.cell(row=row, column=1, value=f"  •  {rec}").font = rec_body_font
            row += 1
        row += 1

    if risk_count_xl["High"] > 0:
        ws_note.cell(row=row, column=1, value="Yüksek Seviye Zafiyetler").font = rec_section_font
        row += 1
        ws_note.cell(row=row, column=1, value=f"{risk_count_xl['High']} adet yüksek seviye zafiyet bulunmaktadır.").font = rec_bold_font
        row += 1
        for rec in [
            "1 hafta içinde yama veya geçici çözüm (workaround) uygulayın.",
            "Güvenlik duvarı kurallarını gözden geçirip sıkılaştırın.",
            "Etkilenen servislerde erişim loglarını analiz edin.",
        ]:
            ws_note.cell(row=row, column=1, value=f"  •  {rec}").font = rec_body_font
            row += 1
        row += 1

    if risk_count_xl["Medium"] > 0:
        ws_note.cell(row=row, column=1, value="Orta Seviye Zafiyetler").font = rec_section_font
        row += 1
        ws_note.cell(row=row, column=1, value=f"{risk_count_xl['Medium']} adet orta seviye zafiyet tespit edilmiştir.").font = rec_bold_font
        row += 1
        for rec in [
            "Aylık bakım penceresinde bu zafiyetleri giderin.",
            "Varsayılan yapılandırmaları kontrol edip sertleştirin.",
        ]:
            ws_note.cell(row=row, column=1, value=f"  •  {rec}").font = rec_body_font
            row += 1
        row += 1

    ws_note.cell(row=row, column=1, value="Genel Güvenlik Önerileri").font = rec_section_font
    row += 1
    for rec in [
        "Düzenli zafiyet taraması yapın — en az ayda bir kez tüm varlıkları tarayın.",
        "Yama yönetimi sürecinizi otomatikleştirin ve SLA'lar tanımlayın.",
        "Ağ segmentasyonu uygulayarak kritik sistemleri izole edin.",
        "Çok faktörlü kimlik doğrulama (MFA) tüm yönetim arayüzlerinde aktif olsun.",
        "Güvenlik bilinci eğitimleri ile personel farkındalığını artırın.",
        "Olay müdahale planınızı güncel tutun ve yılda en az iki kez tatbikat yapın.",
        "Sızma testi (pentest) ile zafiyet taramalarını doğrulayın.",
        "Log yönetimi ile merkezi izleme yapın.",
    ]:
        ws_note.cell(row=row, column=1, value=f"  •  {rec}").font = rec_body_font
        row += 1

    row += 2
    ws_note.merge_cells(f"A{row}:B{row}")
    promo_hdr = ws_note.cell(row=row, column=1, value="SİTEY-VM ile Zafiyet Yönetiminizi Güçlendirin")
    promo_hdr.fill = promo_fill
    promo_hdr.font = promo_font
    promo_hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws_note.row_dimensions[row].height = 28
    row += 2

    for feat in [
        "Çoklu tarayıcı entegrasyonu (OpenVAS, Nessus, Qualys, Acunetix, Nuclei ve dahası)",
        "Yapay zeka destekli zafiyet analizi ve önceliklendirme",
        "Otomatik yama takibi ve SLA yönetimi",
        "Rol tabanlı erişim kontrolü (RBAC) ile ekip yönetimi",
        "İş akışı motoru — zafiyet keşiften çözüme kadar tam izlenebilirlik",
        "Saldırı yüzeyi yönetimi (ASM) ve sürekli varlık keşfi",
        "Kapsamlı raporlama: PDF, Excel, yönetici özeti, teknik detay",
    ]:
        ws_note.cell(row=row, column=1, value=f"  ✓  {feat}").font = promo_body_font
        row += 1

    row += 1
    ws_note.cell(row=row, column=1, value="Web Sitesi").font = rec_bold_font
    ws_note.cell(row=row, column=2, value="https://siteyvm.com").font = Font(color="2563eb", size=11, underline="single")
    row += 1
    ws_note.cell(row=row, column=1, value="İletişim").font = rec_bold_font
    ws_note.cell(row=row, column=2, value="satis@siteyvm.com").font = Font(color="2563eb", size=11, underline="single")

    row += 2
    ws_note.merge_cells(f"A{row}:B{row}")
    warn_cell = ws_note.cell(row=row, column=1, value="⚠  Bu rapor SİTEY-VM Demo Sürümü ile oluşturulmuştur. Kurumsal lisans için: siteyvm.com")
    warn_cell.fill = demo_warn_fill
    warn_cell.font = Font(color="f97316", size=10, italic=True)
    warn_cell.alignment = Alignment(horizontal="center")

    ws_note.column_dimensions["A"].width = 70
    ws_note.column_dimensions["B"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"siteyvm_demo_rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/vuln/")
def list_vulns(
    risk: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    archived: Optional[bool] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Vulnerability)
    if archived is True:
        q = q.filter(Vulnerability.archived == True)
    else:
        q = q.filter((Vulnerability.archived == False) | (Vulnerability.archived == None))
    if risk:
        q = q.filter(Vulnerability.risk == risk)
    if status:
        q = q.filter(Vulnerability.status == status)
    if search:
        pattern = f"%{search}%"
        q = q.filter(
            (Vulnerability.name.ilike(pattern))
            | (Vulnerability.cve.ilike(pattern))
            | (Vulnerability.description.ilike(pattern))
            | (Vulnerability.target_ip.ilike(pattern))
        )
    total = q.count()
    items = q.order_by(Vulnerability.timestamp.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [_vuln_dict(v) for v in items], "total": total, "page": page, "page_size": page_size}


@app.get("/api/vuln/{vuln_id}")
def get_vuln(vuln_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    v = db.query(Vulnerability).filter_by(id=vuln_id).first()
    if not v:
        raise HTTPException(404, "Zafiyet bulunamadı")
    return _vuln_dict(v)


def _vuln_dict(v: Vulnerability) -> dict:
    return {
        "id": v.id,
        "name": v.name or "",
        "agent_uuid": v.agent_uuid or "",
        "scanner": v.scanner or "",
        "cve": v.cve or "",
        "risk": v.risk or "",
        "description": v.description or "",
        "solution": v.solution or "",
        "target_ip": v.target_ip or "",
        "port": v.port or "",
        "service": v.service or "",
        "cvss_score": v.cvss_score or "",
        "status": v.status or "open",
        "timestamp": v.timestamp.isoformat() if v.timestamp else None,
        "archived": bool(v.archived) if v.archived else False,
        "archived_at": v.archived_at.isoformat() if v.archived_at else None,
        "archived_by": v.archived_by or "",
        "has_task": False,
        "task_status": None,
        "task_current_stage": "",
        "task_progress": "",
        "task_assignee": "",
    }


def _top_ips(vulns):
    ip_map = {}
    for v in vulns:
        if v.target_ip:
            ip_map[v.target_ip] = ip_map.get(v.target_ip, 0) + 1
    return [{"ip": k, "count": c} for k, c in sorted(ip_map.items(), key=lambda x: -x[1])[:10]]


@app.post("/api/vuln/manual")
def create_manual_vuln(body: ManualVulnCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    cvss_val = (body.cvss_score or "").strip()
    if cvss_val:
        try:
            cvss_f = float(cvss_val)
            if cvss_f < 0.0 or cvss_f > 10.0:
                raise ValueError
        except ValueError:
            raise HTTPException(400, "CVSS skoru 0.0 ile 10.0 arasında olmalıdır")

    ip_val = (body.target_ip or "").strip()
    if ip_val:
        import ipaddress as _ipa
        try:
            _ipa.ip_address(ip_val)
        except ValueError:
            if not re.match(r'^[\w\.\-\:/]+$', ip_val):
                raise HTTPException(400, "Geçersiz IP adresi formatı")

    port_val = (body.port or "").strip()
    if port_val:
        try:
            port_int = int(port_val)
            if port_int < 0 or port_int > 65535:
                raise ValueError
        except ValueError:
            raise HTTPException(400, "Port 0 ile 65535 arasında bir tam sayı olmalıdır")

    risk_val = body.risk if body.risk in ("Critical", "High", "Medium", "Low", "Info") else "Medium"
    if cvss_val:
        cvss_f = float(cvss_val)
        risk_cvss_map = {
            "Critical": (9.0, 10.0), "High": (7.0, 8.9),
            "Medium": (4.0, 6.9), "Low": (0.1, 3.9), "Info": (0.0, 0.0),
        }
        expected = risk_cvss_map.get(risk_val)
        if expected and not (expected[0] <= cvss_f <= expected[1]) and risk_val != "Info":
            if cvss_f >= 9.0:
                risk_val = "Critical"
            elif cvss_f >= 7.0:
                risk_val = "High"
            elif cvss_f >= 4.0:
                risk_val = "Medium"
            elif cvss_f >= 0.1:
                risk_val = "Low"
            else:
                risk_val = "Info"

    if (body.cve or "").strip() and ip_val and port_val:
        existing = db.query(Vulnerability).filter(
            Vulnerability.cve == _sanitize(body.cve),
            Vulnerability.target_ip == _sanitize(ip_val),
            Vulnerability.port == _sanitize(port_val),
            Vulnerability.archived == False,
        ).first()
        if existing:
            raise HTTPException(409, f"Bu zafiyet zaten mevcut (ID: {existing.id})")

    v = Vulnerability(
        id=str(uuid.uuid4()),
        name=_sanitize(body.name),
        cve=_sanitize(body.cve or ""),
        risk=risk_val,
        description=_sanitize(body.description or ""),
        solution=_sanitize(body.solution or ""),
        target_ip=_sanitize(ip_val),
        port=_sanitize(port_val),
        service=_sanitize(body.service or ""),
        scanner="Manuel",
        cvss_score=_sanitize(cvss_val),
        status="open",
        agent_uuid="manual",
        timestamp=datetime.now(timezone.utc),
    )
    db.add(v)
    db.commit()
    db.refresh(v)
    return _vuln_dict(v)


@app.put("/api/vuln/{vuln_id}")
def update_vuln(vuln_id: str, body: ManualVulnUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    v = db.query(Vulnerability).filter_by(id=vuln_id).first()
    if not v:
        raise HTTPException(404, "Zafiyet bulunamadı")

    cvss_val = (body.cvss_score if body.cvss_score is not None else (v.cvss_score or "")).strip()
    if cvss_val:
        try:
            cvss_f = float(cvss_val)
            if cvss_f < 0.0 or cvss_f > 10.0:
                raise ValueError
        except ValueError:
            raise HTTPException(400, "CVSS skoru 0.0 ile 10.0 arasında olmalıdır")

    ip_val = (body.target_ip if body.target_ip is not None else (v.target_ip or "")).strip()
    if ip_val:
        import ipaddress as _ipa
        try:
            _ipa.ip_address(ip_val)
        except ValueError:
            if not re.match(r'^[\w\.\-\:/]+$', ip_val):
                raise HTTPException(400, "Geçersiz IP adresi formatı")

    port_val = (body.port if body.port is not None else (v.port or "")).strip()
    if port_val:
        try:
            port_int = int(port_val)
            if port_int < 0 or port_int > 65535:
                raise ValueError
        except ValueError:
            raise HTTPException(400, "Port 0 ile 65535 arasında bir tam sayı olmalıdır")

    risk_val = body.risk if body.risk and body.risk in ("Critical", "High", "Medium", "Low", "Info") else (v.risk or "Medium")
    if cvss_val:
        cvss_f = float(cvss_val)
        if body.risk is not None:
            if cvss_f >= 9.0:
                risk_val = "Critical"
            elif cvss_f >= 7.0:
                risk_val = "High"
            elif cvss_f >= 4.0:
                risk_val = "Medium"
            elif cvss_f >= 0.1:
                risk_val = "Low"
            else:
                risk_val = "Info"

    if body.name is not None:
        if not body.name.strip():
            raise HTTPException(400, "Zafiyet adı boş olamaz")
        v.name = _sanitize(body.name)
    if body.cve is not None:
        v.cve = _sanitize(body.cve)
    v.risk = risk_val
    if body.description is not None:
        v.description = _sanitize(body.description)
    if body.solution is not None:
        v.solution = _sanitize(body.solution)
    if body.target_ip is not None:
        v.target_ip = _sanitize(ip_val)
    if body.port is not None:
        v.port = _sanitize(port_val)
    if body.service is not None:
        v.service = _sanitize(body.service)
    if body.cvss_score is not None:
        v.cvss_score = _sanitize(cvss_val)

    db.commit()
    db.refresh(v)
    return _vuln_dict(v)


@app.patch("/api/vuln/{vuln_id}/status")
def update_vuln_status(vuln_id: str, body: VulnStatusUpdate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    valid_statuses = ["open", "in_progress", "resolved", "false_positive"]
    if body.status not in valid_statuses:
        raise HTTPException(400, f"Geçersiz durum. Geçerli değerler: {valid_statuses}")
    v = db.query(Vulnerability).filter_by(id=vuln_id).first()
    if not v:
        raise HTTPException(404, "Zafiyet bulunamadı")
    v.status = body.status
    db.commit()
    return _vuln_dict(v)


@app.delete("/api/vuln/{vuln_id}")
def delete_vuln(vuln_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    v = db.query(Vulnerability).filter_by(id=vuln_id).first()
    if not v:
        raise HTTPException(404, "Zafiyet bulunamadı")
    db.delete(v)
    db.commit()
    return {"ok": True, "message": "Zafiyet silindi"}


@app.post("/api/vuln/bulk_delete")
def bulk_delete_vulns(body: BulkIdsRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Seçilen zafiyetleri toplu sil (maks 100 adet)."""
    if not body.ids:
        raise HTTPException(400, "En az bir zafiyet ID gerekli")
    if len(body.ids) > 100:
        raise HTTPException(400, "Tek seferde en fazla 100 zafiyet silinebilir")
    deleted = db.query(Vulnerability).filter(Vulnerability.id.in_(body.ids)).delete(synchronize_session="fetch")
    db.commit()
    return {"ok": True, "deleted": deleted, "message": f"{deleted} zafiyet silindi"}


@app.post("/api/vuln/bulk_status")
def bulk_update_status(body: BulkStatusRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Seçilen zafiyetlerin durumunu toplu güncelle (maks 100 adet)."""
    valid_statuses = ["open", "in_progress", "resolved", "false_positive"]
    if body.status not in valid_statuses:
        raise HTTPException(400, f"Geçersiz durum. Geçerli değerler: {valid_statuses}")
    if not body.ids:
        raise HTTPException(400, "En az bir zafiyet ID gerekli")
    if len(body.ids) > 100:
        raise HTTPException(400, "Tek seferde en fazla 100 zafiyet güncellenebilir")
    updated = db.query(Vulnerability).filter(Vulnerability.id.in_(body.ids)).update(
        {Vulnerability.status: body.status}, synchronize_session="fetch"
    )
    db.commit()
    return {"ok": True, "updated": updated, "message": f"{updated} zafiyet '{body.status}' olarak güncellendi"}


@app.post("/api/vuln/bulk_false_positive")
def bulk_false_positive(body: BulkIdsRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Seçilen zafiyetleri yanlış pozitif olarak işaretle ve arşivle."""
    if not body.ids:
        raise HTTPException(400, "En az bir zafiyet ID gerekli")
    now = datetime.now(timezone.utc)
    vulns = db.query(Vulnerability).filter(Vulnerability.id.in_(body.ids)).all()
    count = 0
    for v in vulns:
        v.status = "false_positive"
        v.archived = True
        v.archived_at = now
        v.archived_by = user.username
        count += 1
    db.commit()
    return {"ok": True, "updated": count, "message": f"{count} zafiyet yanlış pozitif olarak işaretlendi ve arşivlendi"}


@app.post("/api/vuln/bulk_archive")
def bulk_archive(body: BulkIdsRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Seçilen zafiyetleri arşivle."""
    if not body.ids:
        raise HTTPException(400, "En az bir zafiyet ID gerekli")
    now = datetime.now(timezone.utc)
    vulns = db.query(Vulnerability).filter(Vulnerability.id.in_(body.ids)).all()
    count = 0
    for v in vulns:
        v.archived = True
        v.archived_at = now
        v.archived_by = user.username
        count += 1
    db.commit()
    return {"ok": True, "updated": count, "message": f"{count} zafiyet arşivlendi"}


@app.post("/api/vuln/bulk_unarchive")
def bulk_unarchive(body: BulkIdsRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Seçilen zafiyetleri arşivden çıkar."""
    if not body.ids:
        raise HTTPException(400, "En az bir zafiyet ID gerekli")
    vulns = db.query(Vulnerability).filter(Vulnerability.id.in_(body.ids)).all()
    count = 0
    for v in vulns:
        v.archived = False
        v.archived_at = None
        v.archived_by = None
        count += 1
    db.commit()
    return {"ok": True, "updated": count, "message": f"{count} zafiyet arşivden çıkarıldı"}


@app.post("/api/vuln/archive/{vuln_id}")
def archive_vuln(vuln_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Tek bir zafiyeti arşivle."""
    v = db.query(Vulnerability).filter_by(id=vuln_id).first()
    if not v:
        raise HTTPException(404, "Zafiyet bulunamadı")
    v.archived = True
    v.archived_at = datetime.now(timezone.utc)
    v.archived_by = user.username
    db.commit()
    return _vuln_dict(v)


@app.post("/api/vuln/unarchive/{vuln_id}")
def unarchive_vuln(vuln_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Tek bir zafiyeti arşivden çıkar."""
    v = db.query(Vulnerability).filter_by(id=vuln_id).first()
    if not v:
        raise HTTPException(404, "Zafiyet bulunamadı")
    v.archived = False
    v.archived_at = None
    v.archived_by = None
    db.commit()
    return _vuln_dict(v)


@app.post("/api/scan/import/openvas")
async def import_openvas_xml(file: UploadFile = File(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """OpenVAS XML rapor dosyasını içe aktar.
    ⚠ Information/Log seviye bulgular filtrelenir (import edilmez).
    """
    if not file.filename.endswith(('.xml', '.XML')):
        raise HTTPException(400, "Sadece XML dosyaları kabul edilmektedir")
    
    MAX_UPLOAD_SIZE = 50 * 1024 * 1024
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(413, f"Dosya boyutu çok büyük. Maksimum: {MAX_UPLOAD_SIZE // (1024*1024)}MB")
    
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        raise HTTPException(400, "Geçersiz XML dosyası")
    
    imported = 0
    filtered = 0

    results = root.findall('.//result') or root.findall('.//results/result')
    if not results:
        results = root.findall('.//Result') or root.findall('./result')
    
    for result in results:
        threat = result.findtext('threat') or result.findtext('Threat') or result.findtext('severity') or "Info"
        threat = threat.strip()

        if threat.lower() in ('log', 'information', 'info', 'debug', 'none'):
            filtered += 1
            continue

        risk_map = {
            "High": "High", "Medium": "Medium", "Low": "Low",
            "Alarm": "Critical", "Critical": "Critical",
        }
        risk = risk_map.get(threat, "Low")

        name = result.findtext('name') or result.findtext('Name') or "Bilinmeyen Zafiyet"
        name = name.strip()
        
        host = result.findtext('host') or result.findtext('Host') or ""
        host = host.strip()
        if "<" in host:
            host = host.split("<")[0].strip()
        
        port_text = result.findtext('port') or result.findtext('Port') or ""
        port_text = port_text.strip()
        port_num = port_text.split('/')[0] if port_text else ""
        service = port_text.split('/')[-1] if '/' in port_text else ""

        nvt = result.find('nvt')
        cve = ""
        cvss = ""
        solution = ""
        nvt_family = ""
        description = ""
        impact = ""
        insight = ""

        if nvt is not None:
            cvss_el = nvt.find('cvss_base')
            cvss = cvss_el.text.strip() if cvss_el is not None and cvss_el.text else ""

            refs_el = nvt.find('refs')
            if refs_el is not None:
                for ref in refs_el.findall('ref'):
                    if ref.get('type') == 'cve':
                        cve = ref.get('id', '')
                        break

            sol_el = nvt.find('solution')
            if sol_el is not None and sol_el.text:
                solution = sol_el.text.strip()
                sol_type = sol_el.get('type', '')
                if sol_type:
                    solution = f"[{sol_type}] {solution}"

            family_el = nvt.find('family')
            nvt_family = family_el.text.strip() if family_el is not None and family_el.text else ""

            tags_el = nvt.find('tags')
            if tags_el is not None and tags_el.text:
                tags_text = tags_el.text
                parsed_tags = {}
                for part in tags_text.split('|'):
                    if '=' in part:
                        key, _, val = part.partition('=')
                        parsed_tags[key.strip()] = val.strip()

                if parsed_tags.get('summary'):
                    description = parsed_tags['summary']
                if parsed_tags.get('impact'):
                    impact = parsed_tags['impact']
                if parsed_tags.get('insight'):
                    insight = parsed_tags['insight']
                if not solution and parsed_tags.get('solution'):
                    solution = parsed_tags['solution']
                    if parsed_tags.get('solution_type'):
                        solution = f"[{parsed_tags['solution_type']}] {solution}"

        if not description:
            desc_el = result.find('description') or result.find('Description')
            description = desc_el.text.strip() if desc_el is not None and desc_el.text else ""

        full_desc = description
        if impact:
            full_desc += f"\n\nEtki: {impact}"
        if insight:
            full_desc += f"\n\nDetay: {insight}"

        v = Vulnerability(
            id=str(uuid.uuid4()),
            name=_sanitize(name),
            cve=_sanitize(cve),
            risk=risk,
            description=_sanitize(full_desc[:2000]) if full_desc else "",
            solution=_sanitize(solution[:2000]) if solution else "",
            target_ip=_sanitize(host),
            port=_sanitize(port_num),
            service=_sanitize(service or nvt_family),
            scanner="OpenVAS",
            cvss_score=_sanitize(cvss),
            status="open",
            agent_uuid="openvas-import",
            timestamp=datetime.now(timezone.utc),
        )
        if cve and host and port_num:
            existing = db.query(Vulnerability).filter(
                Vulnerability.cve == _sanitize(cve),
                Vulnerability.target_ip == _sanitize(host),
                Vulnerability.port == _sanitize(port_num),
                Vulnerability.archived == False,
            ).first()
            if existing:
                filtered += 1
                continue
        db.add(v)
        imported += 1
    
    db.commit()
    return {
        "ok": True,
        "imported": imported,
        "filtered": filtered,
        "message": f"{imported} zafiyet içe aktarıldı ({filtered} bilgi seviyesi bulgu filtrelendi)"
    }


ENTERPRISE_MSG = {
    "error": "Bu özellik Kurumsal Sürümde kullanılabilir.",
    "message": "Kurumsal sürüm için iletişime geçin: satis@siteyvm.com | siteyvm.com",
}

@app.get("/api/agent/")
def agents_list(user: User = Depends(get_current_user)):
    return []

@app.get("/api/scan/")
def scans_list(user: User = Depends(get_current_user)):
    return []

@app.get("/api/group/groups")
def groups_list(user: User = Depends(get_current_user)):
    return []

@app.get("/api/network/discovered")
def network_discovered(user: User = Depends(get_current_user)):
    return []

@app.get("/api/network/vlans")
def vlans(user: User = Depends(get_current_user)):
    return []


BLOG_FEED_URL = "https://siteyvm.com/blog-feed.json"
BLOG_CHECK_INTERVAL = 180
BLOG_CACHE_FILE = os.path.join(os.path.dirname(__file__), "blog_cache.json")

_blog_cache = {
    "posts": [],
    "last_check": 0,
    "etag": None,
}


def _load_blog_cache():
    """Disk cache'den blog verilerini yükle."""
    global _blog_cache
    if os.path.exists(BLOG_CACHE_FILE):
        try:
            with open(BLOG_CACHE_FILE, "r", encoding="utf-8") as f:
                _blog_cache = json.load(f)
        except Exception:
            pass


def _save_blog_cache():
    """Blog verilerini disk cache'e kaydet."""
    try:
        with open(BLOG_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(_blog_cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _fetch_blog_feed():
    """siteyvm.com'dan blog feed'ini çek. 3 dakikada 1 kez kontrol eder."""
    global _blog_cache
    now = time.time()

    if now - _blog_cache.get("last_check", 0) < BLOG_CHECK_INTERVAL:
        return _blog_cache.get("posts", [])

    try:
        import urllib.request
        import urllib.error

        req = urllib.request.Request(
            BLOG_FEED_URL,
            headers={"User-Agent": "SITEY-VM Demo/1.3"}
        )
        if _blog_cache.get("etag"):
            req.add_header("If-None-Match", _blog_cache["etag"])

        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode("utf-8"))

        posts = data.get("posts", [])
        _blog_cache["posts"] = posts
        _blog_cache["last_check"] = now
        etag = resp.headers.get("ETag")
        if etag:
            _blog_cache["etag"] = etag
        _save_blog_cache()
        return posts

    except urllib.error.HTTPError as e:
        if e.code == 304:
            _blog_cache["last_check"] = now
            _save_blog_cache()
            return _blog_cache.get("posts", [])
        _blog_cache["last_check"] = now
        if not _blog_cache.get("posts"):
            _blog_cache["posts"] = _get_default_posts()
        _save_blog_cache()
        return _blog_cache.get("posts", [])

    except Exception:
        _blog_cache["last_check"] = now
        if not _blog_cache.get("posts"):
            _blog_cache["posts"] = _get_default_posts()
        _save_blog_cache()
        return _blog_cache.get("posts", [])


def _get_default_posts():
    """Blog feed erişilemezse gösterilecek varsayılan haberler."""
    return [
        {
            "id": "welcome-sitey-vm",
            "title": "SİTEY-VM Zafiyet Yönetim Platformu",
            "summary": "Türkiye'nin en gelişmiş zafiyet yönetim sistemi ile tanışın. 16+ güvenlik aracını tek platformda birleştiren modern çözüm.",
            "url": "https://siteyvm.com",
            "date": "2025-12-01T00:00:00Z",
            "category": "Duyuru",
            "image": None,
        },
        {
            "id": "enterprise-features-2026",
            "title": "Kurumsal Sürüm Yeni Özellikleri",
            "summary": "AI Güvenlik Asistanı, Gelişmiş Raporlama, Agent Yönetimi ve daha fazlası kurumsal sürümde. Detaylar için web sitemizi ziyaret edin.",
            "url": "https://siteyvm.com/#features",
            "date": "2026-01-15T00:00:00Z",
            "category": "Ürün",
            "image": None,
        },
        {
            "id": "pentest-services",
            "title": "Sızma Testi ve Güvenlik Hizmetlerimiz",
            "summary": "Red Team, Blue Team, Purple Team, Pentest, DDoS Testi, Sosyal Mühendislik ve daha fazlası. Profesyonel güvenlik hizmetleri için bizimle iletişime geçin.",
            "url": "https://siteyvm.com/services.html",
            "date": "2026-02-01T00:00:00Z",
            "category": "Hizmetler",
            "image": None,
        },
    ]


_load_blog_cache()


@app.get("/api/notifications/blog")
def get_blog_notifications(user=Depends(get_current_user)):
    """Blog yazılarını bildirim olarak döndür."""
    posts = _fetch_blog_feed()
    notifications = []
    for post in posts:
        notifications.append({
            "id": post.get("id", str(uuid.uuid4())),
            "type": "blog",
            "title": post.get("title", ""),
            "summary": post.get("summary", ""),
            "url": post.get("url", "https://siteyvm.com"),
            "date": post.get("date", ""),
            "category": post.get("category", "Blog"),
            "image": post.get("image"),
        })
    notifications.sort(key=lambda n: n.get("date", ""), reverse=True)
    return {"notifications": notifications}


@app.get("/api/notifications")
def notifications(user=Depends(get_current_user)):
    """Tüm bildirimleri döndür (blog dahil)."""
    posts = _fetch_blog_feed()
    result = []
    for post in posts:
        result.append({
            "id": post.get("id", str(uuid.uuid4())),
            "type": "blog",
            "title": post.get("title", ""),
            "summary": post.get("summary", ""),
            "url": post.get("url", "https://siteyvm.com"),
            "date": post.get("date", ""),
            "category": post.get("category", "Blog"),
            "image": post.get("image"),
        })
    result.sort(key=lambda n: n.get("date", ""), reverse=True)
    return result

@app.get("/api/enterprise/{path:path}")
@app.post("/api/enterprise/{path:path}")
def enterprise_catchall(path: str, user: User = Depends(get_current_user)):
    return ENTERPRISE_MSG


def _find_frontend_dir():
    """Frontend build dizinini bul — hem dev hem EXE modunda."""
    candidates = [
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "build"),
        os.path.join(getattr(sys, '_MEIPASS', ''), "frontend", "build"),
        os.path.join(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__), "frontend", "build"),
    ]
    for d in candidates:
        if os.path.isdir(d) and os.path.isfile(os.path.join(d, "index.html")):
            return d
    return None


frontend_dir = _find_frontend_dir()

if frontend_dir and os.path.isdir(frontend_dir):
    _BLOCKED_PATHS = {
        ".env", ".git", ".gitignore", "debug", "_debug",
        "admin", "config", "env", "settings", "secret",
        "backend", "database.py", "security.py", "models.py",
        ".htaccess", "web.config", "wp-admin", "wp-login",
    }

    @app.get("/{path:path}")
    def serve_frontend(path: str):
        path_lower = path.lower().strip("/")
        first_segment = path_lower.split("/")[0] if path_lower else ""
        if first_segment in _BLOCKED_PATHS or path_lower.startswith("api/"):
            return JSONResponse(status_code=404, content={"detail": "Bulunamadı"})
        safe_path = os.path.normpath(path).lstrip("/").lstrip("\\")
        if ".." in safe_path:
            return JSONResponse(status_code=400, content={"detail": "Geçersiz yol"})
        file_path = os.path.join(frontend_dir, safe_path)
        real_file = os.path.realpath(file_path)
        real_frontend = os.path.realpath(frontend_dir)
        if not real_file.startswith(real_frontend):
            return JSONResponse(status_code=403, content={"detail": "Erişim engellendi"})
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(frontend_dir, "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=5000)
